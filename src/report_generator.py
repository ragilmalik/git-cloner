"""
RagilGitClone - Report Generator Module
Creates XLSX and CSV reports of cloned repositories

Author: Ragilmalik
"""

import csv
from typing import List, Dict, Tuple
from pathlib import Path

def generate_xlsx_report(data: List[Dict], output_path: str) -> Tuple[bool, str]:
    """
    Generate Excel report of cloned repositories

    Returns:
        (success, message)
    """
    try:
        try:
            import pandas as pd
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        except ImportError as e:
            return False, f"Missing dependency: {str(e)}"

        # Filter only successful clones
        successful_clones = [item for item in data if item.get('success', False)]

        if not successful_clones:
            return False, "No successful clones to report"

        # Prepare data for DataFrame
        report_data = []
        for item in successful_clones:
            # Keep raw formatting as requested
            username = str(item.get('username', ''))
            repo_name = str(item.get('repo_name', ''))
            # Remove .git from URL for display
            clean_url = item['url'].replace('.git', '')
            
            report_data.append({
                'Github Repository URL': clean_url,
                'Github Username': username,
                'Github Repository Name': repo_name,
                'Repository Size (in kB)': item['size_kb'],
                'Date & Time of Pull': item['timestamp']
            })

        # Create DataFrame
        df = pd.DataFrame(report_data)

        # Write to Excel with styling
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Clone Report', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Clone Report']

            # Define Styles
            # Table Theme: Dark Calm Gray
            header_fill = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid') # Darker Gray
            row_fill_odd = PatternFill(start_color='2D2D2D', end_color='2D2D2D', fill_type='solid') # Dark Gray
            row_fill_even = PatternFill(start_color='333333', end_color='333333', fill_type='solid') # Slightly Lighter
            
            header_font = Font(name='Segoe UI', size=12, bold=True, color='00FFFF') # Cyan Header Text
            cell_font = Font(name='Segoe UI', size=11, color='FFFFFF') # White Cell Text
            
            thin_border = Border(
                left=Side(style='thin', color='555555'),
                right=Side(style='thin', color='555555'),
                top=Side(style='thin', color='555555'),
                bottom=Side(style='thin', color='555555')
            )
            
            center_align = Alignment(horizontal='center', vertical='center')
            left_align = Alignment(horizontal='left', vertical='center')

            # Apply styles to Header
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = center_align

            # Apply styles to Rows
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                fill = row_fill_even if row_idx % 2 == 0 else row_fill_odd
                for cell in row:
                    cell.fill = fill
                    cell.font = cell_font
                    cell.border = thin_border
                    cell.alignment = left_align

            # Auto-adjust column widths
            for idx, col in enumerate(df.columns):
                max_length = max(
                    df[col].astype(str).map(len).max(),
                    len(col)
                ) + 4
                worksheet.column_dimensions[chr(65 + idx)].width = min(max_length, 60)

        return True, f"XLSX report generated: {output_path}"

    except PermissionError:
        return False, f"Permission denied. Close '{output_path}' if open."
    except Exception as e:
        return False, f"XLSX Error: {str(e)}"


def generate_csv_report(data: List[Dict], output_path: str) -> Tuple[bool, str]:
    """
    Generate CSV report of cloned repositories

    Returns:
        (success, message)
    """
    try:
        # Filter only successful clones
        successful_clones = [item for item in data if item.get('success', False)]

        if not successful_clones:
            return False, "No successful clones to report"

        # Define column headers (exact order)
        headers = [
            'Github Repository URL',
            'Github Username',
            'Github Repository Name',
            'Repository Size (in kB)',
            'Date & Time of Pull'
        ]

        # Write CSV file
        with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=headers)

            # Write header
            writer.writeheader()

            # Write data rows
            for item in successful_clones:
                username = str(item.get('username', ''))
                repo_name = str(item.get('repo_name', ''))
                clean_url = item['url'].replace('.git', '')

                writer.writerow({
                    'Github Repository URL': clean_url,
                    'Github Username': username,
                    'Github Repository Name': repo_name,
                    'Repository Size (in kB)': item['size_kb'],
                    'Date & Time of Pull': item['timestamp']
                })

        return True, f"CSV report generated: {output_path}"

    except PermissionError:
        return False, f"Permission denied. Close '{output_path}' if open."
    except Exception as e:
        return False, f"CSV Error: {str(e)}"


def generate_report(data: List[Dict], output_path: str, format_type: str = 'xlsx') -> Tuple[bool, str]:
    """
    Generate report in specified format

    Returns:
        (success, message)
    """
    # Add appropriate extension
    if format_type.lower() == 'xlsx':
        if not output_path.endswith('.xlsx'):
            output_path += '.xlsx'
        return generate_xlsx_report(data, output_path)
    elif format_type.lower() == 'csv':
        if not output_path.endswith('.csv'):
            output_path += '.csv'
        return generate_csv_report(data, output_path)
    else:
        return False, f"Unknown format: {format_type}"


# Testing
if __name__ == "__main__":
    # Sample test data
    test_data = [
        {
            'success': True,
            'url': 'https://github.com/torvalds/linux.git',
            'username': 'torvalds',
            'repo_name': 'linux',
            'size_kb': 1234567.89,
            'timestamp': '29-11-2025 14:30:45'
        },
        {
            'success': True,
            'url': 'https://github.com/microsoft/vscode.git',
            'username': 'microsoft',
            'repo_name': 'vscode',
            'size_kb': 234567.12,
            'timestamp': '29-11-2025 14:35:20'
        },
        {
            'success': False,
            'url': 'https://github.com/invalid/repo.git',
            'username': 'invalid',
            'repo_name': 'repo',
            'size_kb': 0,
            'timestamp': '29-11-2025 14:40:00',
            'error': 'Repository not found'
        }
    ]

    print("Testing Report Generation:")
    print("\n1. CSV Report:")
    generate_csv_report(test_data, "test_report.csv")

    print("\n2. XLSX Report:")
    generate_xlsx_report(test_data, "test_report.xlsx")